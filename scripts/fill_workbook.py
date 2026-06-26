import json
import re
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment


LABEL_SHEET_PATTERN = re.compile(r"^(\d+)#\(\d+\)$")


def cbm_formula(row: int) -> str:
    dimension = f'SUBSTITUTE(G{row},"×","*")'
    first_star = f'FIND("*",{dimension})'
    second_star_text = f'SUBSTITUTE({dimension},"*","@",2)'
    second_star = f'FIND("@",{second_star_text})'
    length = f"VALUE(LEFT({dimension},{first_star}-1))"
    width = f"VALUE(MID({dimension},{first_star}+1,{second_star}-{first_star}-1))"
    height = f"VALUE(RIGHT({dimension},LEN({dimension})-{second_star}))"
    return f'=IFERROR(E{row}*{length}*{width}*{height}/1000000000,"")'


def ref(cell: str) -> str:
    return f"'FJPN送货单'!{cell}"


def blank_ref(cell: str) -> str:
    source = ref(cell)
    return f'=IF({source}="","",{source})'


def prune_label_sheets(workbook, keep_count: int) -> None:
    for sheet_name in list(workbook.sheetnames):
        match = LABEL_SHEET_PATTERN.match(sheet_name)
        if match and int(match.group(1)) > keep_count:
            del workbook[sheet_name]


def english_label_font_size(text: str) -> int:
    length = len(text or "")
    if length <= 16:
        return 36
    if length <= 22:
        return 30
    if length <= 30:
        return 24
    if length <= 40:
        return 20
    if length <= 52:
        return 17
    return 15


def fit_english_label(sheet, english_name: str) -> None:
    cell = sheet["F7"]
    font = copy(cell.font)
    font.sz = english_label_font_size(english_name)
    cell.font = font
    cell.alignment = Alignment(
        horizontal="center",
        vertical="center",
        shrink_to_fit=True,
        wrap_text=False,
    )


def repair_label_formulas(workbook, items: list[dict]) -> None:
    keep_count = len(items)
    for index in range(1, keep_count + 1):
        row = index + 6
        sheet_name = f"{index}#({row})"
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        sheet["C2"] = blank_ref("C49")
        sheet["D5"] = blank_ref("C3")
        sheet["D7"] = blank_ref("C4")
        sheet["F7"] = blank_ref(f"D{row}")
        fit_english_label(sheet, items[index - 1].get("englishName", ""))
        sheet["D9"] = blank_ref("C2")
        sheet["F10"] = blank_ref(f"C{row}")
        sheet["D11"] = blank_ref("I3")
        sheet["D13"] = blank_ref("C5")
        sheet["D15"] = blank_ref("I2")
        sheet["D17"] = blank_ref(f"A{row}")
        sheet["H17"] = blank_ref(f"I{row}")
        sheet["D19"] = f"={ref('E41')}"
        sheet["H19"] = blank_ref(f"G{row}")
        sheet["B22"] = blank_ref("C48")
        sheet["B23"] = blank_ref("C47")
        sheet["B24"] = blank_ref("C46")


def fill_workbook(template_path: Path, output_path: Path, data: dict) -> None:
    workbook = load_workbook(template_path)
    items = data.get("items", [])[:33]
    prune_label_sheets(workbook, len(items))

    sheet = workbook["FJPN送货单"]
    sheet["C2"] = data.get("elevatorSpec", "")
    sheet["C3"] = data.get("projectName") or data.get("consignee", "")
    sheet["C4"] = data.get("address", "")
    sheet["C5"] = data.get("factoryNumber", "")
    sheet["I2"] = data.get("shipDate", "")
    sheet["I3"] = data.get("contractNumber", "")
    sheet["I4"] = data.get("contactPhone", "")
    sheet["I5"] = data.get("recipient", "")

    header_style = copy(sheet["G6"]._style)
    if "G6:H6" in [str(merged_range) for merged_range in sheet.merged_cells.ranges]:
        sheet.unmerge_cells("G6:H6")
    sheet["G6"] = "尺寸"
    sheet["H6"] = "CBM"
    sheet["G6"]._style = copy(header_style)
    sheet["H6"]._style = copy(header_style)

    def set_cell(address: str, value) -> None:
        cell = sheet[address]
        if isinstance(cell, MergedCell):
            return
        cell.value = value

    for offset in range(33):
        row = offset + 7
        item = items[offset] if offset < len(items) else None
        if item:
            set_cell(f"A{row}", item.get("actualBoxNo") or f"{offset + 1}#")
            set_cell(f"B{row}", item.get("boxNo") or offset + 1)
            set_cell(f"C{row}", item.get("chineseName", ""))
            set_cell(f"D{row}", item.get("englishName", ""))
            set_cell(f"E{row}", item.get("quantity") or 1)
            set_cell(f"F{row}", item.get("unit") or "箱")
            set_cell(f"G{row}", item.get("size", ""))
            set_cell(f"I{row}", item.get("weight") or "/")
            set_cell(f"J{row}", item.get("note") or None)
        else:
            for column in "ABCDEFGHJ":
                set_cell(f"{column}{row}", None)
            set_cell(f"I{row}", "/")
        sheet[f"H{row}"] = cbm_formula(row)
        sheet[f"H{row}"].number_format = "0.000"

    sheet["H40"] = None
    sheet["H41"] = "=SUM(H7:H39)"
    sheet["H41"].number_format = "0.000"

    repair_label_formulas(workbook, items)
    workbook.save(output_path)


def main() -> int:
    if len(sys.argv) != 4:
        print("Usage: fill_workbook.py <template.xlsx> <data.json> <output.xlsx>", file=sys.stderr)
        return 2

    template_path = Path(sys.argv[1])
    data_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])
    data = json.loads(data_path.read_text(encoding="utf-8"))
    fill_workbook(template_path, output_path, data)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
