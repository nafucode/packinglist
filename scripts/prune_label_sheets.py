import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: prune_label_sheets.py <xlsx_path> <keep_count>", file=sys.stderr)
        return 2

    xlsx_path = Path(sys.argv[1])
    keep_count = max(0, int(sys.argv[2]))
    label_sheet_pattern = re.compile(r"^(\d+)#\(\d+\)$")

    workbook = load_workbook(xlsx_path)
    removed = []

    for sheet_name in list(workbook.sheetnames):
        match = label_sheet_pattern.match(sheet_name)
        if not match:
            continue
        label_no = int(match.group(1))
        if label_no > keep_count:
            removed.append(sheet_name)
            del workbook[sheet_name]

    workbook.save(xlsx_path)
    print(",".join(removed))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
